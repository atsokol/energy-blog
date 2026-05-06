#!/usr/bin/env python3
"""Read all LinkedIn analytics Excel files from analytics/ and print JSON."""
import json
import sys
import subprocess
from pathlib import Path

try:
    import openpyxl
except ImportError:
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"], check=True)
    import openpyxl

analytics_dir = Path("analytics")
if not analytics_dir.exists():
    print(json.dumps({"error": "analytics/ directory not found"}))
    sys.exit(1)

results = []

for xlsx_file in sorted(analytics_dir.glob("*.xlsx")):
    wb = openpyxl.load_workbook(xlsx_file, data_only=True)
    record = {"file": xlsx_file.name}

    # PERFORMANCE sheet: read by label in column A (row positions vary between files)
    if "PERFORMANCE" in wb.sheetnames:
        ws = wb["PERFORMANCE"]
        for row in ws.iter_rows(values_only=True):
            if row[0] is not None:
                label = str(row[0]).strip()
                value = row[1] if len(row) > 1 else None
                if label:
                    record[label] = value

    # TOP DEMOGRAPHICS sheet
    if "TOP DEMOGRAPHICS" in wb.sheetnames:
        ws = wb["TOP DEMOGRAPHICS"]
        demo = []
        headers = None
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = [str(c).strip() if c else "" for c in row]
            else:
                if any(c is not None for c in row):
                    demo.append(dict(zip(headers, row)))
        record["demographics"] = demo

    wb.close()
    results.append(record)

print(json.dumps(results, default=str, indent=2))
